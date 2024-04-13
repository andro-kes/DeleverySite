from typing import Any
from django.db.models.query import QuerySet
from django.forms import BaseModelForm
from django.shortcuts import render, redirect 
from django.http import HttpResponse, JsonResponse
from django.views.generic.base import View
from django.views.generic.detail import DetailView
from .forms import SearchProductForm, CreateOrderForm, ChangeStatusForm
from .models import OrderModel, UserCartModel
from django.views.generic.list import ListView
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
import heapq
import os
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .filters import OrderFilter
from django.views.generic.edit import CreateView
from accounts.models import Cities
import random
import re

class MainView(View):
    """ 
    Класс для функционала главной страницы
    """
    template_name = 'main/main.html'
    
    def get(self, request):
        data = {
            'form': SearchProductForm(),
        }
        return render(request, self.template_name, data)
    def post(self, request, *args, **kwargs):
        form = SearchProductForm(request.POST)
        if form.is_valid():
            value = form.cleaned_data['search_field']
        # находим товары, содержащие данную категорию
        
        queryset = OrderModel.objects.filter(category__icontains=value)
        # получаем айди всех найденных элементов
        request.session['queryset_ids'] = list(queryset.values_list('id', flat=True))
        request.session['search_word'] = value
        return redirect('results')

class CreateOrderView(CreateView):
    template_name = 'main/create.html'
    form_class = CreateOrderForm
    success_url = 'main/orders_company.html'
    
    @classmethod
    def get_number(self):
        # Генерируем случайную строку из 12 цифр
        random_number = ''.join(str(random.randint(0, 9)) for _ in range(12))

        # Форматируем строку в формате ****-****-****
        formatted_number = re.sub(r'(\d{4})(\d{4})(\d{4})', r'\1-\2-\3', random_number)

        return formatted_number
    
    def form_valid(self, form: BaseModelForm):
        form.instance.company = self.request.user
        form.instance.picture1 = self.request.FILES.get('picture1')
        form.instance.picture2 = self.request.FILES.get('picture2')
        form.instance.picture3 = self.request.FILES.get('picture3')
        form.instance.number = self.get_number()
        form.save()
        return super().form_valid(form)

    
    def get_context_data(self, **kwargs: Any):
        context =  super().get_context_data(**kwargs)
        context['search_form'] = SearchProductForm()
        context['cities'] = Cities.objects.all()
        return context

class ResultsView(ListView):
    '''
    Класс для показа товаров по результатам поиска
    '''
    template_name = 'main/results.html'
    filterset_class = OrderFilter

    def get_queryset(self):
        queryset_ids = self.request.session.get('queryset_ids', [])
        return OrderModel.objects.filter(id__in=queryset_ids).order_by('price')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        page = self.request.GET.get('page')
        queryset = self.get_queryset()
        filterset = self.filterset_class(self.request.GET, queryset=queryset)
        paginator = Paginator(filterset.qs, 10)  # количество объектов на странице
        try:
            objects = paginator.page(page)
        except PageNotAnInteger:
            objects = paginator.page(1)
        except EmptyPage:
            objects = paginator.page(paginator.num_pages)
        context['page_obj'] = objects
        context['form'] = SearchProductForm()
        context['search_word'] = self.request.session.get('search_word')
        context['filterset'] = filterset
        return context

@login_required 
def add_to_cart(request):
    """
    Добавление в корзину
    """
    product_id = request.POST.get('product_id')
    print(product_id)
    if product_id:
        product = OrderModel.objects.get(id=product_id)
        user_cart = UserCartModel.objects.get(user=request.user)
        user_cart.products.add(product)
        return JsonResponse({'status': 'ok'})
    else:
        return JsonResponse({'status': 'error'})

class ProductDetailView(DetailView):
    '''
    Класс для отдельного представления продукта
    '''
    model = OrderModel
    template_name = 'main/detail.html'
    context_object_name = 'product'
    
    @classmethod
    def dijkstra(self, graph, start):
        distances = {city: float('infinity') for city in graph}
        distances[start] = 0
        pq = [(0, start)]

        while pq:
            current_distance, current_city = heapq.heappop(pq)

            if current_distance > distances[current_city]:
                continue

            for neighbor in graph[current_city]:
                distance = current_distance + neighbor['distance']

                if distance < distances[neighbor['city']]:
                    distances[neighbor['city']] = distance
                    heapq.heappush(pq, (distance, neighbor['city']))

        return distances
    
    @classmethod
    def to_graph(self):
        # Открываем Excel файл
        app_path = os.path.dirname(__file__)

        # Построение пути к файлу econom
        file_path_earth = os.path.join(app_path, 'Расстояния_econom.xlsx')
        
        # fast
        file_path_air = os.path.join(app_path, 'Расстояния_fast.xlsx')
        
        # Выбираем активный листы econom
        wb = load_workbook(file_path_earth)
        sheet = wb.active
        # fast
        wb_air = load_workbook(file_path_air)
        sheet_air = wb_air.active
        
        # Преобразуем данные в виде графа
        # Получаем список всех городов, чтобы потом использовать их индексы
        cities_index_list = []
        for city in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for name in city:
                if name:
                    cities_index_list.append(name)
                    
        # Создаем словарь для econom
        data_list_econom = []
        for city in sheet.iter_rows(min_row=2, values_only=True):
            data = [{'city': city[0]}]
            for i in range(1, len(city)):
                if city[i] != None:
                    if city[i] != '-':
                        data.append({'city': cities_index_list[i-1], 'distance': int(city[i])})
            data_list_econom.append(data)
        
        # fast    
        data_list_fast = []
        for city in sheet_air.iter_rows(min_row=2, values_only=True):
            data = [{'city': city[0]}]
            for i in range(1, len(city)):
                if city[i] != None:
                    if city[i] != '-':
                        data.append({'city': cities_index_list[i-1], 'distance': int(city[i])})
            data_list_fast.append(data)
            
        # Преобразуем данные в виде графа
        #econom
        graph_econom = {}
        for city_data in data_list_econom:
            city = city_data[0]['city']
            graph_econom[city] = [{'city': data['city'], 'distance': data['distance']} for data in city_data[1:]]
        
        # fast
        graph_fast = {}
        for city_data in data_list_fast:
            city = city_data[0]['city']
            graph_fast[city] = [{'city': data['city'], 'distance': data['distance']} for data in city_data[1:]]
        # Возвращаем оба графа
        return (graph_econom, graph_fast)
    
    def get_context_data(self, **kwargs: Any):
        # Скорость доставки в км/ч
        v_econom = 50
        v_fast = 200
        
        # Цена за каждый час в рублях
        price_list_econom = 100
        price_list_fast = 500
        
        # Получаем контекст от супер класса 
        context =  super().get_context_data(**kwargs) 
         
        # Получаем город пользователя
        print(self.request.user.address.all())
        address = self.request.user.address.all()[0].name
        
        # Получим склады товара
        list_warehouse = []
        for warehouse in self.object.list_warehouse.all():
            list_warehouse.append(warehouse.name)
            
        # Получаем пункты выдачи компании
        company = self.object.company
        list_pick_up_point = []
        for pick_up_point in company.list_pick_up_point.all():
            list_pick_up_point.append(pick_up_point.name)
            
        # Получим все расстояния
        
        # econom
        shortest_paths_econom = self.dijkstra(self.to_graph()[0], address)
        #fast
        shortest_paths_fast = self.dijkstra(self.to_graph()[1], address)
        
        # Найдем расстояния до каждого склада
        list_distance_econom = []
        for warehouse in list_warehouse:
            distance = shortest_paths_econom[warehouse]
            list_distance_econom.append(distance)
        # fast
        list_distance_fast = []
        for warehouse in list_warehouse:
            distance = shortest_paths_fast[warehouse]
            list_distance_fast.append(distance)
        
        # Проверям, есть ли в городе заказчика пункт выдачи компании-производителя
        if address in list_pick_up_point:
            # Время для econom
            context['econom'] = min(list_distance_econom)/v_econom
            # Цена econom 
            context['result_price_econom'] = (min(list_distance_econom)/v_econom)*price_list_econom + int(self.object.price)
            
            # Время для fast
            context['fast'] = min(list_distance_fast)/v_fast
            # Цена fast
            context['result_price_fast'] = (min(list_distance_fast)/v_fast)*price_list_fast + int(self.object.price)
        # Если пункта нет  
        else:
            # Время для econom
            context['econom'] = (min(list_distance_econom)/v_econom)
            # Цена econom 
            context['result_price_econom'] = (min(list_distance_econom)/v_econom)*2*price_list_econom+ int(self.object.price)
            
            # Время для fast
            context['fast'] = (min(list_distance_fast)/v_fast)
            # Цена fast
            context['result_price_fast'] = (min(list_distance_fast)/v_fast)*2*price_list_fast + int(self.object.price)
            
            # Уведомление о том, что пунктов выдачи нет
            context['notifiction'] = 'В вашем городе нет нашего пункта выдачи, поэтому доставка дороже :('
            
        # Если город, пункт выдачи и склад - один и тот же город, то доставка не требуется
        if int(context['econom']) == 0:
            context['notifiction'] = 'Пункт выдачи в вашем городе готов выдать вам заказ'
            context['econom'] = 'Готов к выдаче'
            context['result_price_econom'] = 'Готов к выдаче'
            context['fast'] = 'Готов к выдаче'
            context['result_price_fast'] = 'Готов к выдаче'
            
        context['form'] = SearchProductForm()
        return context
    
class OrdersListView(ListView):
    template_name = 'main/orders_company.html'
    
    def get_queryset(self):
        return OrderModel.objects.filter(company=self.request.user)
    
    def get_context_data(self, **kwargs: Any):
        context = super().get_context_data(**kwargs)
        context['orders'] = self.get_queryset()
        context['form'] = ChangeStatusForm()
        return context
    
    def post(self, request, *args, **kwargs):
        form = ChangeStatusForm(request.POST)
        if form.is_valid():
            status = form.cleaned_data['status']
            number = form.cleaned_data['number']
            order = OrderModel.objects.get(number=number)
            order.status = status
            order.save()
        return redirect('orders')
    
def privacy_policy(request):
    return render(request, 'main/privacy_policy.html')